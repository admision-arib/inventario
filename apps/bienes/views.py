from django.db.models import Max
from django.http import JsonResponse
from django.shortcuts import render, redirect,get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.utils import timezone
from django.core.exceptions import ValidationError
from django.http import HttpResponse

from datetime import date

from .models.sede import Sede
from .models.movimientos import MovimientoBien
from .models import Siga
from .models.bien import Bien
from .models.area import Area
from apps.usuarios.models import Usuario
from .forms import SedeForm
from .forms import MovimientoBienForm
from .forms import BienForm
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from apps.usuarios.decorators import (
    gestionar_bienes_requerido,
    ver_reportes_requerido,
)
from django.core.exceptions import PermissionDenied

# =============================
# GENERADOR DE CÓDIGO PATRIMONIAL
# =============================

def generar_codigo_patrimonial():
    año = date.today().year
    prefijo = f"IESTP-{año}-"
    ultimo = Bien.objects.filter(codigo_patrimonial__startswith=prefijo).aggregate(
        max_codigo=Max('codigo_patrimonial')
    )['max_codigo']
    if ultimo:
        try:
            secuencia = int(ultimo.split('-')[-1]) + 1
        except (ValueError, IndexError):
            secuencia = 1
    else:
        secuencia = 1
    return f"{prefijo}{secuencia:05d}"


# =============================
# ALTA DE BIEN
# =============================
@login_required
@gestionar_bienes_requerido
def crear_bien(request):

    if request.method == 'POST':

        print("===== DEBUG POST =====")
        for key, value in request.POST.items():
            print(f"{key}: {value}")
        print("=======================")

        form = BienForm(request.POST)

        if form.is_valid():

            try:
                with transaction.atomic():

                    bien = form.save(commit=False)

                    # ✅ VALIDACIÓN COMPLETA DEL MODELO
                    bien.full_clean()

                    bien.save()

                messages.success(request, "✅ Bien registrado correctamente")
                return redirect('lista_bienes')

            except Exception as e:
                print("ERROR:", e)
                messages.error(request, f" Error al guardar: {str(e)}")

        else:
            print(form.errors)
            messages.error(request, " Revise los campos del formulario")

    else:
        form = BienForm()

    return render(request, 'bienes/form.html', {
        'form': form
    })

# =============================
# LISTA DE BIENES (CON FILTRO POR ROL)
# =============================
@login_required
def lista_bienes(request):
    user = request.user
    # Determinar alcance según rol
    if user.puede_gestionar_bienes() or user.es_auditor() or user.es_consulta():
        bienes = Bien.objects.all().order_by('-fecha_creacion')
    elif user.es_responsable_bien() or user.es_inventariador():
        bienes = Bien.objects.filter(usuario_responsable=user).order_by('-fecha_creacion')
    else:
        # Por seguridad, si no tiene ningún rol válido, denegamos
        raise PermissionDenied("No tiene permisos para ver la lista de bienes.")

    return render(request, 'bienes/lista.html', {'bienes': bienes})


# =============================
# SEDES (solo gestión de bienes)
# =============================
@login_required
@gestionar_bienes_requerido
def crear_sede(request):
    if request.method == 'POST':
        form = SedeForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Sede creada exitosamente.')
            return redirect('lista_sedes')
    else:
        form = SedeForm()
    return render(request, 'sedes/form.html', {'form': form, 'accion': 'Crear'})


@login_required
@gestionar_bienes_requerido
def lista_sedes(request):
    sedes = Sede.objects.all().order_by('codigo')
    return render(request, 'sedes/lista.html', {'sedes': sedes})


@login_required
@gestionar_bienes_requerido
def editar_sede(request, pk):
    sede = get_object_or_404(Sede, pk=pk)
    if request.method == 'POST':
        form = SedeForm(request.POST, instance=sede)
        if form.is_valid():
            form.save()
            messages.success(request, 'Sede actualizada.')
            return redirect('lista_sedes')
    else:
        form = SedeForm(instance=sede)
    return render(request, 'sedes/form.html', {'form': form, 'accion': 'Editar', 'sede': sede})

@login_required
@gestionar_bienes_requerido
def desactivar_sede(request, pk):
    sede = get_object_or_404(Sede, pk=pk)
    sede.activo = False
    sede.save()
    messages.warning(request, f'Sede {sede.nombre} desactivada.')
    return redirect('lista_sedes')


# =============================
# BUSCADOR CATÁLOGO (acceso ampliado)
# =============================
@login_required
def buscar_catalogo(request):
    user = request.user
    # Permitir a ADMIN, CONTROL PATRIMONIAL, AUDITOR, CONSULTA y RESPONSABLE BIEN
    if not (user.puede_gestionar_bienes() or
            user.es_auditor() or
            user.es_consulta() or
            user.es_responsable_bien()):
        raise PermissionDenied("No tiene permiso para consultar el catálogo.")
    q = request.GET.get('q', '')
    items = Siga.objects.filter(denominacion__icontains=q)[:10]
    data = [
        {
            'id': i.id,
            'codigo': i.codigo_siga,
            'nombre': i.denominacion
        }
        for i in items
    ]
    return JsonResponse(data, safe=False)


# =============================
# USCADOR DE BIENES (para movimientos)
# =============================
@login_required
@gestionar_bienes_requerido
def buscar_bien(request):
    q = request.GET.get('q', '')
    bienes = Bien.objects.filter(
        denominacion__icontains=q
    ) | Bien.objects.filter(
        codigo_patrimonial__icontains=q
    )
    bienes = bienes.distinct()[:10]
    data = [
        {
            'id': b.id,
            'codigo': b.codigo_patrimonial,
            'denominacion': b.denominacion
        }
        for b in bienes
    ]
    return JsonResponse(data, safe=False)


# =============================
# MOVIMIENTOS
# =============================
@login_required
@gestionar_bienes_requerido
def lista_movimientos(request):
    movimientos = MovimientoBien.objects.select_related(
        'bien', 'sede_origen', 'sede_destino', 'usuario_origen', 'usuario_destino', 'registrado_por'
    ).all().order_by('-fecha_movimiento')
    return render(request, 'movimientos/lista.html', {'movimientos': movimientos})


@login_required
@gestionar_bienes_requerido
def crear_movimiento(request):
    if request.method == 'POST':
        form = MovimientoBienForm(request.POST)

        if form.is_valid():
            with transaction.atomic():
                movimiento = form.save(commit=False)
                movimiento.registrado_por = request.user
                movimiento.fecha_movimiento = timezone.now()
                movimiento.save()

                bien = movimiento.bien

                # Lógica según tipo
                if movimiento.tipo_movimiento == 'TRANSFERENCIA':
                    bien.sede = movimiento.sede_destino
                    bien.usuario_responsable = None

                elif movimiento.tipo_movimiento == 'ASIGNACION':
                    bien.usuario_responsable = movimiento.usuario_destino

                elif movimiento.tipo_movimiento == 'DEVOLUCION':
                    bien.sede = movimiento.sede_destino
                    bien.usuario_responsable = None

                # PRÉSTAMO no cambia estado permanente

                bien.save()

            messages.success(request, f'Movimiento registrado para {bien}')
            return redirect('lista_movimientos')

    else:
        form = MovimientoBienForm()

    return render(request, 'movimientos/form.html', {'form': form})


# =============================
# REPORTES
# =============================
@login_required
@ver_reportes_requerido
def reportes(request):
    return render(request, 'reportes.html')


# =============================
# IMPORTACIÓN Y EXPORTACIÓN EXCEL
# =============================
@login_required
@gestionar_bienes_requerido
def importar_bienes_excel(request):
    if request.method != 'POST' or not request.FILES.get('archivo_excel'):
        messages.error(request, " No se proporcionó ningún archivo válido.")
        return redirect('lista_bienes')

    excel_file = request.FILES['archivo_excel']

    if not excel_file.name.endswith(('.xlsx', '.xlsm')):
        messages.error(request, ' Formato inválido. Debe subir un archivo .xlsx o .xlsm')
        return redirect('lista_bienes')

    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        sheet = wb.active

        # 1. MAPEADO DINÁMICO DE COLUMNAS: Lee la fila 1 y detecta en qué columna está cada dato
        header_map = {}
        for col_idx, cell in enumerate(sheet[1], start=1):
            if cell.value:
                # Almacena el nombre en minúsculas y sin espacios para evitar errores de tipeo
                header_map[str(cell.value).strip().lower()] = col_idx

        # Validamos que al menos existan las dos columnas esenciales de identificación
        if 'codigo_siga' not in header_map and 'denominacion' not in header_map:
            messages.error(request,
                           " Estructura inválida. El Excel debe contener al menos las columnas 'codigo_siga' o 'denominacion'.")
            return redirect('lista_bienes')

        errores = []
        bienes_creados_count = 0
        mapa_estados = {'BUENO': 'B', 'REGULAR': 'R', 'MALO': 'M', 'INSERVIBLE': 'I', 'NUEVO': 'N'}

        # Datos maestros base de Sede y Área
        sede_obj = Sede.objects.first()
        area_obj = Area.objects.first()
        usuario_defecto = Usuario.objects.filter(is_active=True).first() or request.user

        if not sede_obj or not area_obj:
            messages.error(request,
                           " Configuración incompleto: Registre una Sede y un Área en el sistema antes de continuar.")
            return redirect('lista_bienes')

        with transaction.atomic():
            # Iterar desde la fila 2 (omitir cabeceras)
            for row_idx in range(2, sheet.max_row + 1):
                # Validar si la fila está completamente en blanco
                row_values = [sheet.cell(row=row_idx, column=c).value for c in range(1, sheet.max_column + 1)]
                if not any(row_values):
                    continue

                # Función auxiliar para leer los datos de la celda según su nombre de columna mapeado
                def get_val(column_name):
                    col_pos = header_map.get(column_name.lower())
                    if col_pos:
                        val = sheet.cell(row=row_idx, column=col_pos).value
                        return str(val).strip() if val is not None else ""
                    return ""

                # --- EXTRACCIÓN DE DATOS DE LA FILA ---
                codigo_siga_raw = get_val('codigo_siga')
                denominacion = get_val('denominacion')
                marca = get_val('marca')
                modelo = get_val('modelo')
                num_serie = get_val('serie') if get_val('serie') else get_val('numero_serie')
                estado_raw = get_val('estado').upper()

                num_doc = get_val('numero_documento') or "MIGRACION"
                nro_pecosa = get_val('nro_pecosa') or "PEC-MIGRA"

                # Conversión segura de valor de documento a flotante numérico
                valor_doc_raw = get_val('valor_documento')
                try:
                    # Limpiamos símbolos comunes de moneda si existieran en la celda de texto
                    if valor_doc_raw:
                        valor_doc_raw = valor_doc_raw.replace('S/.', '').replace(',', '').strip()
                    valor_doc = float(valor_doc_raw) if valor_doc_raw else 0.00
                except ValueError:
                    valor_doc = 0.00

                # --- CONTROL DE DUPLICADOS ---
                if num_serie and num_serie != "" and Bien.objects.filter(numero_serie=num_serie).exists():
                    continue

                # --- PROCESAMIENTO DINÁMICO DE USUARIOS (DNI O LOGIN) ---
                # Leemos lo que venga en el Excel para el campo del responsable
                dni_responsable = get_val('usuario_responsable')
                dni_asignado = get_val('usuario_asignado')

                # Buscamos el usuario por su número de documento en la BD, si no existe usamos el por defecto
                resp_obj = Usuario.objects.filter(num_documento=dni_responsable).first() if dni_responsable else None
                asig_obj = Usuario.objects.filter(num_documento=dni_asignado).first() if dni_asignado else None

                if not resp_obj: resp_obj = usuario_defecto
                if not asig_obj: asig_obj = usuario_defecto

                # --- BÚSQUEDA ROBUSTA EN CATÁLOGO SIGA ---
                catalogo = None
                if codigo_siga_raw:
                    catalogo = Siga.objects.filter(
                        codigo_siga=codigo_siga_raw).first() or Siga.objects.filter(codigo=codigo_siga_raw).first()

                if not catalogo and denominacion:
                    catalogo = Siga.objects.filter(denominacion__icontains=denominacion).first()

                if not catalogo and denominacion:
                    palabras = denominacion.split()
                    if palabras:
                        catalogo = Siga.objects.filter(denominacion__icontains=palabras[0]).first()

                if not catalogo:
                    errores.append(
                        f"Fila {row_idx}: No se localizó concordancia en el Catálogo SIGA para '{denominacion}'.")
                    continue

                # --- CONTROL DE FECHAS COMPATIBLES ---
                # Extraemos la fecha del excel o generamos un bloque seguro coherente de migración
                fecha_pecosa_raw = get_val('fecha_salida_pecosa')

                if fecha_pecosa_raw:
                    try:
                        # Si viene como objeto datetime desde openpyxl
                        cell_date = sheet.cell(row=row_idx, column=header_map['fecha_salida_pecosa']).value
                        if isinstance(cell_date, (datetime, date)):
                            fecha_pecosa = cell_date if isinstance(cell_date, date) else cell_date.date()
                        else:
                            fecha_pecosa = datetime.strptime(fecha_pecosa_raw, "%d/%m/%Y").date()
                    except Exception:
                        fecha_pecosa = date.today()
                else:
                    fecha_pecosa = date.today()

                # La fecha de compra siempre la forzamos a ser 2 días anterior a la de salida de PECOSA
                fecha_compra = fecha_pecosa - timedelta(days=2)
                estado_id = mapa_estados.get(estado_raw, 'B')

                try:
                    # Instanciamos el objeto mapeado
                    bien = Bien(
                        catalogo=catalogo,
                        denominacion=denominacion if denominacion else catalogo.denominacion,
                        tipo_doc_adquisicion='031',
                        numero_documento=num_doc,
                        valor_documento=valor_doc,
                        fecha_documento=fecha_compra,
                        nro_pecosa=nro_pecosa,
                        fecha_salida_pecosa=fecha_pecosa,
                        sede=sede_obj,
                        area=area_obj,
                        usuario_responsable=resp_obj,
                        usuario_asignado=asig_obj,
                        marca=marca,
                        modelo=modelo,
                        numero_serie=num_serie,
                        estado_conservacion=estado_id,
                        observaciones="Importado vía actualización masiva de inventario."
                    )

                    bien.full_clean()
                    bien.save()
                    bienes_creados_count += 1

                except ValidationError as e:
                    err_msg = ", ".join([f"{k}: {v}" for k, v in e.message_dict.items()])
                    errores.append(f"Fila {row_idx} (Error de Validación): {err_msg}")
                except Exception as ex:
                    errores.append(f"Fila {row_idx} (Error Crítico): {str(ex)}")

            if errores:
                transaction.set_rollback(True)
                return render(request, 'bienes/importar_errores.html', {'errores': errores})

        if bienes_creados_count > 0:
            messages.success(request,
                             f" ¡Éxito Total! Se procesó el archivo correctamente. {bienes_creados_count} bienes sincronizados en la plataforma.")
        else:
            messages.warning(request, "⚠️ No se detectaron nuevas filas con números de serie únicos para procesar.")

    except Exception as e:
        messages.error(request, f" Error general al procesar la lectura del archivo: {str(e)}")

    return redirect('lista_bienes')


@login_required
@gestionar_bienes_requerido
def exportar_bienes_excel(request):
    """
    Genera un archivo Excel en tiempo real con TODOS los bienes registrados
    en la base de datos, sirviendo como reporte y como plantilla de importación.
    """
    # 1. Crear el libro de Excel en memoria
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bienes_Registrados"

    # Forzar a que se vean las líneas de cuadrícula en Excel
    ws.views.sheetView[0].showGridLines = True

    # 2. Estilos visuales (Color Azul Corporativo para que combine con tu software SIPAT)
    HEADER_FILL = PatternFill(start_color="2A31C1", end_color="2A31C1", fill_type="solid")
    WHITE_FONT = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
    DARK_FONT = Font(name="Segoe UI", size=10, color="333333")

    thin_line = Side(border_style="thin", color="D1D5DB")
    border_cell = Border(left=thin_line, right=thin_line, top=thin_line, bottom=thin_line)

    # 3. Cabeceras EXACTAS que tu importador necesita leer en la Fila 1
    headers = [
        "codigo_siga",
        "denominacion",
        "marca",
        "modelo",
        "serie",
        "estado",
        "numero_documento",
        "valor_documento",
        "nro_pecosa"
    ]

    # Escribir la fila de cabeceras
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border_cell
    ws.row_dimensions[1].height = 28

    # 4. CONSULTA REAL A LA BASE DE DATOS: Trae todos tus bienes cargados
    bienes_en_bd = Bien.objects.all().select_related('catalogo')

    # Diccionario inverso para transformar la inicial de la BD al texto que entiende el Excel
    conversor_estados = {
        'B': 'Bueno', 'R': 'Regular', 'M': 'Malo',
        'I': 'Inservible', 'N': 'Nuevo'
    }

    # Escribir cada bien real de la base de datos en una fila del Excel
    for row_idx, bien in enumerate(bienes_en_bd, start=2):

        # Mapeamos los atributos de tu modelo Django a las columnas del Excel
        row_data = [
            bien.catalogo.codigo_siga if bien.catalogo else "",  # Trae el código real del catálogo asociado
            bien.denominacion,
            bien.marca or "",
            bien.modelo or "",
            bien.numero_serie or "",
            conversor_estados.get(bien.estado_conservacion, "Bueno"),  # Pasa 'B' -> 'Bueno'
            bien.numero_documento or "",
            bien.valor_documento or 0.00,
            bien.nro_pecosa or ""
        ]

        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DARK_FONT
            cell.border = border_cell

            # Formatear celdas según su tipo de dato para que se vea ordenado
            if col_idx in [1, 5, 6, 7, 9]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col_idx == 8:  # Columna de dinero (Monto)
                cell.alignment = Alignment(horizontal="right", vertical="center")
                cell.number_format = '"S/." #,##0.00'  # Formato de Moneda Peruana
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

        ws.row_dimensions[row_idx].height = 20

    # 5. Ajustar automáticamente el ancho de las columnas según el texto más largo
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max(max_len + 5, 15)

    # 6. Configurar la respuesta HTTP para obligar al navegador a descargar el archivo XLSX
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="Inventario_Actual_SIPAT.xlsx"'

    wb.save(response)
    return response

from django.views.decorators.http import require_POST
@require_POST
@login_required
@gestionar_bienes_requerido
def crear_item_catalogo(request):
    codigo = request.POST.get('codigo', '').strip()
    denominacion = request.POST.get('denominacion', '').strip()

    if not codigo or not denominacion:
        return JsonResponse({'success': False, 'error': 'Código y denominación son obligatorios.'}, status=400)

    # Verificar si ya existe
    if Siga.objects.filter(codigo_siga=codigo).exists():
        return JsonResponse({'success': False, 'error': 'Ya existe un ítem con ese código.'}, status=400)

    item = Siga.objects.create(
        codigo_siga=codigo,
        denominacion=denominacion
    )

    return JsonResponse({
        'success': True,
        'item': {
            'id': item.id,
            'codigo': item.codigo_siga,
            'nombre': item.denominacion
        }
    })


@login_required
def obtener_info_area(request, area_id):
    area = get_object_or_404(Area, pk=area_id)
    jefe = area.jefe
    data = {
        'jefe_id': jefe.pk if jefe else None,
        'jefe_nombre': jefe.nombre_completo if jefe else '',
        'usuarios': list(
            Usuario.objects.filter(area=area, is_active=True).values('id', 'first_name', 'last_name')
        )
    }
    return JsonResponse(data)


@login_required
def datos_origen_bien(request, bien_id):
    bien = get_object_or_404(Bien, pk=bien_id)
    data = {
        'sede_origen_id': bien.sede_id,
        'area_origen_id': bien.area_id if bien.area else None,
        'usuario_origen_id': bien.usuario_responsable_id,
    }
    return JsonResponse(data)